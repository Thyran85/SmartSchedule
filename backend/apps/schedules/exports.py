import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse

from .models import Cours, ScheduleVersion

DAYS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
TIME_SLOTS = [
    ('07:00', '10:00'),
    ('10:00', '10:15', 'PAUSE'),
    ('10:15', '12:00'),
    ('12:00', '14:00', 'PAUSE_DÉJEUNER'),
    ('14:00', '16:00'),
    ('16:00', '16:15', 'PAUSE'),
    ('16:15', '18:00'),
]


def build_grid(courses, entity_type='classe', entity_id=None):
    """Build a 2D grid: rows=time slots, cols=days."""
    grid = {}
    for day_idx, day in enumerate(DAYS):
        for slot_label, slot_end, *rest in TIME_SLOTS:
            if rest and rest[0] in ('PAUSE', 'PAUSE_DÉJEUNER'):
                grid[(day_idx, slot_label)] = rest[0]
                continue
            grid[(day_idx, slot_label)] = None

    for c in courses:
        key = (c.jour_semaine, c.heure_debut.strftime('%H:%M'))
        grid[key] = c

    return grid, TIME_SLOTS


def export_schedule_excel(request, entity_type, entity_id, version_id=None):
    """Export schedule as Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Emploi du temps"

    courses = _get_courses(entity_type, entity_id, version_id)
    if not courses:
        return None

    entity_name = _get_entity_name(entity_type, entity_id)
    ws.cell(row=1, column=1, value=f"Emploi du temps - {entity_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Headers
    ws.cell(row=3, column=1, value="Créneau")
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=1).font = header_font

    for i, day in enumerate(DAYS):
        cell = ws.cell(row=3, column=i + 2, value=day)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    grid, time_slots = build_grid(courses)
    row = 4
    for slot_label, slot_end, *rest in TIME_SLOTS:
        ws.cell(row=row, column=1, value=slot_label)
        if rest and rest[0] in ('PAUSE', 'PAUSE_DÉJEUNER'):
            for i in range(len(DAYS)):
                cell = ws.cell(row=row, column=i + 2, value=rest[0])
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        else:
            for day_idx in range(len(DAYS)):
                course = grid.get((day_idx, slot_label))
                if course:
                    if entity_type == 'classe':
                        value = f"{course.matiere.nom}\n{course.enseignant}\n{course.salle.nom if course.salle else ''}"
                    elif entity_type == 'enseignant':
                        value = f"{course.classe.nom}\n{course.matiere.nom}\n{course.salle.nom if course.salle else ''}"
                    else:
                        value = f"{course.classe.nom}\n{course.matiere.nom}\n{course.enseignant}"
                    ws.cell(row=row, column=day_idx + 2, value=value)
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    for i in range(len(DAYS)):
        ws.column_dimensions[chr(66 + i)].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{entity_type}_{entity_id}.xlsx"'
    return response


def export_schedule_pdf(request, entity_type, entity_id, version_id=None):
    """Export schedule as PDF file."""
    courses = _get_courses(entity_type, entity_id, version_id)
    if not courses:
        return None

    entity_name = _get_entity_name(entity_type, entity_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"Emploi du temps - {entity_name}",
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        f"Emploi du temps - {entity_name}",
        styles['Title']
    ))
    elements.append(Spacer(1, 12))

    # Build table data
    grid, time_slots = build_grid(courses)
    table_data = [['Créneau'] + DAYS]

    for slot_label, slot_end, *rest in TIME_SLOTS:
        row_data = [slot_label]
        if rest and rest[0] in ('PAUSE', 'PAUSE_DÉJEUNER'):
            for _ in range(len(DAYS)):
                row_data.append(rest[0])
        else:
            for day_idx in range(len(DAYS)):
                course = grid.get((day_idx, slot_label))
                if course:
                    if entity_type == 'classe':
                        text = f"{course.matiere.nom}<br/>{course.enseignant}"
                        if course.salle:
                            text += f"<br/>{course.salle.nom}"
                    elif entity_type == 'enseignant':
                        text = f"{course.classe.nom}<br/>{course.matiere.nom}"
                    else:
                        text = f"{course.classe.nom}<br/>{course.matiere.nom}<br/>{course.enseignant}"
                    row_data.append(text)
                else:
                    row_data.append('')
        table_data.append(row_data)

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{entity_type}_{entity_id}.pdf"'
    return response


def _get_courses(entity_type, entity_id, version_id=None):
    """Get courses filtered by entity type."""
    if version_id:
        qs = Cours.objects.filter(version_id=version_id)
    else:
        active = ScheduleVersion.objects.filter(est_active=True).first()
        if not active:
            return []
        qs = Cours.objects.filter(version=active)

    qs = qs.select_related('matiere', 'enseignant', 'salle', 'classe')

    filters = {
        'classe': 'classe_id',
        'enseignant': 'enseignant_id',
        'salle': 'salle_id',
    }
    filter_field = filters.get(entity_type)
    if filter_field:
        qs = qs.filter(**{filter_field: entity_id})

    return qs


def _get_entity_name(entity_type, entity_id):
    """Get the display name for an entity."""
    from ..classes.models import Classe
    from ..teachers.models import Enseignant
    from ..rooms.models import Salle

    models = {
        'classe': Classe,
        'enseignant': Enseignant,
        'salle': Salle,
    }
    model = models.get(entity_type)
    if model:
        try:
            obj = model.objects.get(id=entity_id)
            return str(obj)
        except model.DoesNotExist:
            pass
    return f"{entity_type} #{entity_id}"
